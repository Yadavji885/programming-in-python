import openpyxl
import random

cl = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789"
pc = "1234567890"
workbook = openpyxl.load_workbook("Aadhar.xlsx")
Aadhar = workbook["Aadhar"]

class AadharDetails:

    def Get_Aadhar(self):
        
        rows = Aadhar.max_row
        cols = Aadhar.max_column
        frn = input("Enter your First Name :\n")
        lsn = input("Enter your Last Name :\n")
        dob = input("Enter your date of birth (like 27/5/2024):\n")
        Ad = input("Enter your Address with pincode :\n")
        fn = input("Enter your Father's name :\n")
        mn = input("Enter you Mother's name :\n")
        pn = int(input("Enter your phone number :\n"))
        l = [0,0,frn,lsn,dob,Ad,fn,mn,pn]
        otp = ""

        for i in range(1,7):

            otp = otp + random.choice(pc)
        with open("UIDAI Phone confirmation.txt","w") as f:
            text = f.write(f"This ia an automated message by UIDAI.\nThis is sent to you because you are trying to upload your data on UIDAI.\nHere is your otp : {otp}.\nThis is only valid upto 2 minutes or until you entered this correctly.\nYou have only 3 tries with this otp.\nThank you for using our website and uploading your data.")
        otpc = input("Enter the otp you recieved on your phone :\n")

        if otp == otpc:

            with open("UIDAI Phone confirmation.txt","w") as f:
                text = f.write("Thank you for joining us.\nYou have succesfully logged in to our website.")
        else:

            with open("UIDAI Phone confirmation.txt","w") as f:
                text = f.write(f"This ia an automated message by UIDAI.\nThis is sent to you because you are trying to upload your data on UIDAI.\nHere is your otp : {otp}.\nThis is only valid upto 2 minutes or until you entered this correctly.\nYou have only 3 tries with this otp.\nThank you for using our website and uploading your data.")
            otp = input("You entered the wrong otp please fill correct otp :\n")

        for a in range(2,cols+1):
            Aadhar.cell(rows+1,a).value = l[a]
        an = ""

        for b in range(13):

            an = an + random.choice(pc)
        Aadhar.cell(rows+1,1).value = int(an)

    def Details_check(self):
        
        rows = Aadhar.max_row
        cols = Aadhar.max_column
        an = input("Enter your aadhar no. :\n")

        while len(an) != 12:

            an = input("Please enter valid aadhar no. :\n")
        capatcha = ""
        
        for i in range(0,6):

            capatcha = capatcha + random.choice(cl)
        print("\n",capatcha)
        c = input("\nEnter the capatcha :\n")

        while c != capatcha:

            capatcha = ""

            for i in range(0,6):
                capatcha = capatcha + random.choice(cl)
            print(capatcha)

            c = input("\nYou entered wrong capatcha, please enter the correct capatcha :\n")
        
        for a in range(1,rows+1):

            for b in range(1,cols+1):

                if Aadhar.cell(a,b).value == int(an):

                    for c in range(1,cols+1):

                        print(f"{Aadhar.cell(b,c).value} : {Aadhar.cell(a,c).value}")

    def Update(self):
        rows = Aadhar.max_row
        cols = Aadhar.max_column
        o = {
            0 : "Aadhar Number",
            1 : "First Name",
            2 : "Last Name",
            3 : "D.O.B.",
            4 : "Address",
            5 : "Father's Name",
            6 : "Mother's Name",
            7 : "Phone No."
        }
        
        oa = int(input("\nWhat do you want to update?\n1) First Name\n2) Last Name\n3) D.O.B.\n4) Address\n5) Father's Name\n6) Mother's Name\n7) Phone No.\n\nEnter your choice : "))
        c = o.get(oa)
        coln = 0
        for a in range(1,cols+1):
            if Aadhar.cell(1,a).value == c:
                coln = Aadhar.cell(1,a).column
        an = input("\nEnter your aadhar no. :\n")
        while len(an) != 12:

            an = input("\nPlease enter valid aadhar no. :\n")
        rn = 0
        t = False
        for b in range(1,rows+1):

            if Aadhar.cell(b,1).value == int(an):

                t =True
                rn = Aadhar.cell(b,1).row

        if t == True:

            pass

        else:

            print("You are not registered on our website. Please register yourself.")
            ask = input("Do you want to register now or after sometime? [y/n]: ")

            if ask == "y":

                print("\n")
                self.Get_Aadhar()

            else :

                print("Thank you for your visit and please don't forget to get registered here as it is mandatory for every citizen of this country")
                exit()
        
        capatcha = ""
        
        for j in range(0,6):

            capatcha = capatcha + random.choice(cl)

        print("\n",capatcha)

        c = input("\nEnter the capatcha :\n")

        while c != capatcha:

            capatcha = ""

            for i in range(0,6):

                capatcha = capatcha + random.choice(cl)

            print("\n",capatcha)

            c = input("\nYou entered wrong capatcha, please enter the correct capatcha :\n")

        otp = ""

        for i in range(1,7):

            otp = otp + random.choice(pc)

        with open("UIDAI Phone confirmation.txt","w") as f:

            text = f.write(f"This ia an automated message by UIDAI.\nThis is sent to you because you are trying to update your data on UIDAI.\nHere is your otp : {otp}.\nThis is only valid upto 2 minutes or until you entered this correctly.\nYou have only 3 tries with this otp.\nThank you for using our website and uploading your data.")
        otpc = input("Enter the otp you recieved on your phone :\n")

        while otpc != otp:

            otp = input("You entered the wrong otp please fill correct otp :\n")
            with open("UIDAI Phone confirmation.txt","w") as f:

                text = f.write(f"This ia an automated message by UIDAI.\nThis is sent to you because you are trying to update your data on UIDAI.\nHere is your otp : {otp}.\nThis is only valid upto 2 minutes or until you entered this correctly.\nYou have only 3 tries with this otp.\nThank you for using our website and uploading your data.")
        
        if otp == otpc:

            print("You have succesfully logged in.")
            with open("UIDAI Phone confirmation.txt","w") as f:

                text = f.write("Thank you for joining us.\nYou have succesfully logged in to our website.")
        
        
        ud = input(f"\nEnter your {c} :\n")

        temp = Aadhar.cell(rn,coln).value

        Aadhar.cell(rn,coln).value = ud 

        for b in range(1,cols+1):

            print(f"{o.get(b-1)} : {Aadhar.cell(rn,b).value}")

        confirm = input("\nAre you sure that these details are filled correctly ?[y,n]\n")

        while confirm != "y":

            Aadhar.cell(rn,coln).value = temp

            oa = int(input("\nWhat do you want to update?\n1) First Name\n2) Last Name\n3) D.O.B.\n4) Address\n5) Father's Name\n6) Mother's Name\n7) Phone No.\n\nEnter your choice : "))
            
            c = o.get(oa)

            coln = 0

            for a in range(1,cols+1):

                if Aadhar.cell(1,a).value == c:

                    coln = Aadhar.cell(1,a).column

            ud = input(f"Enter your {c} :\n")

            temp = Aadhar.cell(rn,coln).value

            Aadhar.cell(rn,coln).value = ud
            
            for b in range(1,cols+1):

                print(f"{o.get(b-1)} : {Aadhar.cell(rn,b).value}")

            confirm = input("\nAre you sure that these details are filled correctly ?[y,n]\n")

            if confirm == "y":

                print("\nYour details will be updated soon,Thank you you for using our website.\n")

                workbook.save("Aadhar.xlsx")

            elif confirm == "n":

                Aadhar.cell(rn,coln).value = temp


a = AadharDetails()

a.Get_Aadhar()

a.Details_check()

a.Update()

workbook.save("Aadhar.xlsx")
