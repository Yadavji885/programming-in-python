# importing every modules to be used in the project

from tabulate import tabulate
from datetime import datetime , timedelta
import mysql.connector
import openpyxl
import random
# connecting mysql database to python
mydb = mysql.connector.connect(host = "localhost",user = "root",passwd = "Jatin09*",database = "Train_Reservation_System")
cur = mydb.cursor()

# connecting excel spreadsheet with python for train data
workbook1 = openpyxl.load_workbook("Trains_data.xlsx")
Trains = workbook1["trains"]
rows = Trains.max_row
col = Trains.max_column

# q1 = "create table Train_Data(train_no int,train_name varchar(50),Start varchar(20),End varchar(20),Start_stop_no int,End_stop_no int,Distance int,Departure_time time,Arrival_time time,Start_date date,End_date date);"
# cur.execute(q1)

# inserting train data into mysql table

# for i in range(2,rows+1):
#     train = []
#     for j in range(1,col+1):
        
#         train = train + [Trains.cell(i,j).value]
    
    # q2 = "insert into Train_Data (train_no ,train_name ,Start ,End ,Start_stop_no ,End_stop_no ,Distance ,Departure_time ,Arrival_time ,Start_date,End_date) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
    # cur.execute(q2,train)


# q3 = "create table Users_data(user_id int primary key,user_name varchar(30),password varchar(20))"
# cur.execute(q3)

# connecting excel spreadsheet for user data and inserting the data in mysql table

# workbook2 = openpyxl.load_workbook("Users.xlsx")
# User = workbook2["IRCTC_Users"]
# for a in range(2,(User.max_row)+1):
#     u = []
#     for b in range(1,(User.max_column)+1):
#         u += [User.cell(a,b).value]

    # q = "insert into Users_data (user_id,user_name,password) values (%s,%s,%s)"
    # cur.execute(q,u)
 
# creating a table named ticket to store imp data of the ticket for the official records

# q4 = "create table Ticket(PNR int primary key,train_id int,user_id int,No_of_passengers int,departing_from varchar(30),arriving_to varchar(30),fare int,distance int,Booking_time varchar(25),Departure_time time,Arrival_time time,Departure_date date,Arrival_date date);"
# cur.execute(q4)

# creating a table named ticket_info to store important data of the ticket to be provided to user

# q5 = "create table ticket_info(PNR int,Train_id int,passenger_no int,passenger_name varchar(20),Age int,Gender varchar(10),Seat_type varchar(30),seat_no int,Coach_no varchar(3),Status varchar(10),Departing_from varchar(30),Arriving_to varchar(30),Booking_Time time,Start_Time time,End_Time time,Start_date date,End_date date);"
# cur.execute(q5)


class IRCTC:

    # defining some variable which will be used in every function in class IRCTC

    def __init__(self):
        self.user_id = None
        self.user_name = None
    
    # Greeting the user

    def Welcome(self):
        print("\n**************************\n     Welcome to IRCTC\n**************************","\n")

    # Signing in using the users table in mysql database

    def Signin(self):
        
        global c
        u_id = int(input("Enter your user id (If you don't have one, enter any random number):\n"))
        q = f'select user_name from Users_data where user_id = {u_id}'
        cur.execute(q)
        w1 = cur.fetchone()
        if w1:
            y = w1[0]
            q1 = f'select password from Users_data where user_id = {u_id}'
            cur.execute(q1)
            w = cur.fetchone()
            x = w[0]
            u_pas = input("Enter your password :\n")
            if u_pas == x:
                print(f"**************************           Welcome {y} To IRCTC           **************************")
                self.user_name = y
                self.user_id = u_id
                c = True
            else:
                v = 0
                while u_pas != x:
                    if v < 2:
                        print("You entered wrong password\n")
                        u_pas = input("Enter correct password :\n")
                        if u_pas == x:
                            print(f"**************************           Welcome {y} To IRCTC           **************************")
                            self.user_name = y
                            self.user_id = u_id
                            c = True
                            break
                        else:
                            v += 1
                    else:
                        print("You entered wrong password three times try again after some time")
                        c = False
                        break


        else:
            print("There is no user signed up with this id")
            r = input("Do you want to sign up or not [y for yes,press anything for no] :\n")
            if r == "y":
                user_name = input("Enter your first name only :\n")
                password = input("Enter your password :\n")
                q3 = "select user_id from users_data order by user_id desc limit 1"
                cur.execute(q3)
                t = cur.fetchone()
                user_id = t[0] + 1
                q2 = "insert into users_data values (%s,%s,%s)"
                cur.execute(q2,(user_id,user_name,password))
                print(f"Your user_id is {user_id}")
                self.user_name = user_name
                self.user_id = user_id
                print("Thankyou for signing up with us")
                mydb.commit()
                c = True
            else:
                print("\n","**************************           Thank you for visiting IRCTC           **************************","\n")
                c = False
        
    # Function to book a ticket using the train_data table in mysql database

    def Booking(self):
        cd = input("From which city do you want to travel :\n")
        ca = input("To which city do you want to travel :\n")
        cd = cd.capitalize()
        ca = ca.capitalize()
        q = f"select train_no,train_name,Start,End,Start_stop_no,End_stop_no,Distance,Departure_time,Arrival_time,TIMESTAMP(CURDATE() + INTERVAL 1 DAY, departure_time) AS Start_date,TIMESTAMP(DATE_ADD(CURDATE(), INTERVAL (1 + IF(arrival_time <= departure_time, 1, 0)) DAY),arrival_time) AS End_date from train_data where Start = %s and End = %s;"
        cur.execute(q,(cd,ca))
        a = cur.fetchall()
        
        # if there is any train from the given source to destination
        
        if a:
            columns = [desc[0] for desc in cur.description] 
            print(tabulate(a, headers=columns, tablefmt="psql"))
            l = len(a)
            if l > 1:
                t = int(input("Which train you want to go with ?\n"))
                q6 = "select * from train_data where train_no = %s and Start = %s and End = %s;"
                cur.execute(q6,[t,cd,ca])
                a = cur.fetchall()
                inp = int(input(f"Do you want to book ticket in {a[0][1]} for tomorrow [1 for yes and 0 for no] :\n"))
            else:
                inp = int(input(f"Do you want to book ticket in {a[0][1]} for tomorrow [1 for yes and 0 for no] :\n"))
            
            # if user wants to book ticket in the train

            if inp == 1 :

                inpu = int(input("Do you want to book ticket for yourself or someone else [1 for yourself and 2 for others]:\n"))

                # if user wants to book ticket for himself/herself or for someone else

                if inpu == 1 :

                    # for himself/herself

                    train_id = a[0][0]
                    
                    user_id = self.user_id
                    user_name = self.user_name

                    sp = int(input("Mention your coach preference [1 for First tier AC,2 for Second tier AC,3 for Third tier AC] :\n"))
                    coach = None
                    seat = None
                    if sp == 1:
                        seat = "First tier AC"
                        c = "A"
                    elif sp == 2:
                        seat = "Second tier AC"
                        c = "B"
                    elif sp == 3:
                        seat = "Third tier AC"
                        c = "C"
                    age = int(input("Enter your age :\n"))
                    gender = input("Enter your gender [m for male and f for female] :\n")
                    gender = gender.capitalize()
                    time = datetime.now().strftime("%Y/%m/%d %H:%M")
                    distance = a[0][6]
                    if c == "A":
                        fare = distance * 3
                    elif c == "B":
                        fare = distance * 2
                    elif c == "C":
                        fare = distance * 1.5
                    PNR = random.randint(1000,2000)
                    seat_no = random.randint(1,90)
                    coach_no = random.randint(1,10)
                    coach = f"{c}{coach_no}"
                    dt = (datetime.min + a[0][7]).time()
                    at = (datetime.min + a[0][8]).time()
                    tomorrow = datetime.now().date() + timedelta(days=1)
                    dd = datetime.combine(tomorrow, dt)
                    if at <= dt:
                        ad = tomorrow + timedelta(days=1)
                    else:
                        ad = tomorrow
                    ad = datetime.combine(ad, at)
                    
                    otp = random.randint(100000,999999)
                    with open("messages/OTP.txt","w") as f:
                        f.write(f"Your OTP is {otp}")

                    entered_otp = int(input(f"An OTP has been sent for the payement of {fare}, please enter the OTP to proceed further :\n"))

                    while True:
                        if  entered_otp == otp:
                            print("Payement Done successfully")
                            with open("messages/OTP.txt","w") as g:
                                g.write("Your OTP is ******")
                                g.flush()
                            break
                        else:
                            print("You entered wrong OTP")
                            entered_otp = int(input("Please enter correct OTP to proceed further :\n"))

                    q5 = "insert into ticket(PNR,train_id,user_id,No_of_passengers,departing_from,arriving_to,fare,distance,booking_time,Departure_time,Arrival_time,Departure_date,Arrival_date) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                    l = [PNR,train_id,user_id,1,cd,ca,fare,distance,time,dt,at,dd,ad]
                    cur.execute(q5,l)
                    q6 = "insert into ticket_info(PNR,Train_id,passenger_no,passenger_name,Age,Gender,Seat_type,seat_no,Coach_no,Status,Departing_from,Arriving_to,Booking_Time,Start_Time,End_Time,Start_date,End_date) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                    l1 = [PNR,train_id,1,user_name,age,gender,seat,seat_no,coach,"Confirmed",cd,ca,time,dt,at,dd,ad]
                    cur.execute(q6,l1)
                    print(f"\nThanks for visiting our service your PNR is {PNR}\n")

                    P = PNR
                    q8 = "select * from ticket_info where PNR = %s"
                    cur.execute(q8,(P,))
                    r = cur.fetchall()
                    columns = [desc[0] for desc in cur.description] 
                    print(tabulate(r, headers=columns, tablefmt="psql"))

                else:

                    # for someone else

                    r = int(input("How many tickets you want to book :\n"))
                    w = 1
                    coach = None
                    seat = None
                    sp = int(input("Mention your coach preference [1 for First tier AC,2 for Second tier AC,3 for Third tier AC] :\n"))
                    if sp == 1:
                        seat = "First tier AC"
                        c = "A"
                    elif sp == 2:
                        seat = "Second tier AC"
                        c = "B"
                    elif sp == 3:
                        seat = "Third tier AC"
                        c = "C"
                    while w <= r:
                        train_id = a[0][0]
                        user_id = self.user_id
                        name = input(f"Enter passenger {w} name :\n")
                        age = int(input(f"Enter passenger {w} age :\n"))
                        gender = input("Enter your gender [m for male and f for female] :\n")
                        gender = gender.capitalize()
                        time = datetime.now().strftime("%Y/%m/%d %H:%M")
                        distance = a[0][6]
                        if c == "A":
                            fare = distance * 3 * r
                        elif c == "B":
                            fare = distance * 2 * r
                        elif c == "C":
                            fare = distance * 1.5 * r
                        if w == 1:
                            seat_no = random.randint(1,90)
                            coach_no = random.randint(1,10)
                            coach = f"{c}{coach_no}"
                        else:
                            seat_no += 1
                            coach = coach
                        dt = (datetime.min + a[0][7]).time()
                        at = (datetime.min + a[0][8]).time()
                        tomorrow = datetime.now().date() + timedelta(days=1)
                        dd = datetime.combine(tomorrow, dt)
                        if at <= dt:
                            ad = tomorrow + timedelta(days=1)
                        else:
                            ad = tomorrow
                        ad = datetime.combine(ad, at)
                        if w == 1:
                            PNR = random.randint(1000,2000)
                            q7 = "select PNR from ticket where PNR = %s"
                            cur.execute(q7,(PNR,))
                            e = cur.fetchone()
                            while e:
                                PNR = random.randint(1000, 2000)
                                cur.execute("SELECT PNR FROM ticket WHERE PNR = %s", (PNR,))
                                e = cur.fetchone()
                            

                            q5 = "insert into ticket(PNR,train_id,user_id,No_of_passengers,departing_from,arriving_to,fare,distance,booking_time,Departure_time,Arrival_time,Departure_date,Arrival_date) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                            l = [PNR,train_id,user_id,r,cd,ca,fare,distance,time,dt,at,dd,ad]
                            cur.execute(q5,l)
                        else:
                            pass
                        q6 = "insert into ticket_info(PNR,Train_id,passenger_no,passenger_name,Age,Gender,Seat_type,seat_no,Coach_no,Status,Departing_from,Arriving_to,Booking_Time,Start_Time,End_Time,Start_date,End_date) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                        l1 = [PNR,train_id,w,name,age,gender,seat,seat_no,coach,"confirmed",cd,ca,time,dt,at,dd,ad]
                        cur.execute(q6,l1)
                        
                        w += 1
                    otp = random.randint(100000,999999)
                    with open("messages/OTP.txt","w") as f:
                        f.write(f"Your OTP is {otp}")

                    entered_otp = int(input(f"An OTP has been sent for the payement of {fare}, please enter the OTP to proceed further :\n"))

                    while True:
                        if  entered_otp == otp:
                            print("Payement Done successfully")
                            with open("messages/OTP.txt","w") as g:
                                g.write("Your OTP is ******")
                                g.flush()
                            break
                        else:
                            print("You entered wrong OTP")
                            entered_otp = int(input("Please enter correct OTP to proceed further :\n"))
                    
                    print(f"\nThanks for visiting our service your PNR is {PNR}\n")

                    P = PNR
                    q8 = "select * from ticket_info where PNR = %s"
                    cur.execute(q8,(P,))
                    r = cur.fetchall()
                    columns = [desc[0] for desc in cur.description] 
                    print(tabulate(r, headers=columns, tablefmt="psql"))

        # if there is no train from the given source to destination

        else:
            print(f"\nSorry we dont have any trains from {cd} to {ca}")

    # Function to search the ticket using PNR number

    def Searching(self):
        user_id = self.user_id
        print("You can search your ticket using PNR number")
        P = input("what's your PNR ?\n")
        q15 = "select PNR from ticket where user_id = %s and PNR = %s"
        cur.execute(q15,(user_id,P))
        m = cur.fetchall()
        if not m:
            print(f"There is no ticket booked with this user id having PNR number {P}")
        else:
            print("Here is your ticket information")            
            q8 = "select * from ticket_info where PNR = %s"
            cur.execute(q8,(P,))
            r = cur.fetchall()
            columns = [desc[0] for desc in cur.description] 
            print(tabulate(r, headers=columns, tablefmt="psql"))

    # Function to cancel the ticket of any person using PNR number

    def cancellation(self):
        user_id = self.user_id
        P = int(input("what's your PNR ?\n"))
        q7 = "select PNR from ticket where PNR = %s and user_id = %s"
        cur.execute(q7,(P,user_id))
        if cur.fetchone() is None:
            print(f"There is no ticket booked with this user id having PNR number {P}")
        else:
            print("Here is your ticket information")
            q8 = "select * from ticket_info where PNR = %s"
            cur.execute(q8,(P,))
            r = cur.fetchall()
            columns = [desc[0] for desc in cur.description] 
            print(tabulate(r, headers=columns, tablefmt="psql"))
            q12 = "select No_of_passengers from ticket where PNR = %s"
            q13 = "select Status from ticket_info where PNR = %s"
            cur.execute(q13,(P,))
            l = cur.fetchall()
            cur.execute(q12,(P,))
            o = cur.fetchall()
            if o[0][0] == 1 and l[0][0] == "confirmed":
                x = input("Are you sure you want to cancel the ticket [y for yes and n for no] :\n")
                if x == "y":
                    q11 = "update ticket_info set Status = %s where PNR = %s"
                    l1 = ["Cancelled",P]
                    cur.execute(q11,l1)
                    print("Your ticket has been cancelled successfully")
            elif o[0][0] > 1 and l[0][0] == "confirmed":
                n = int(input("Enter the no of passenger's whose tickets you want to cancel :\n"))
                if n == o[0][0]:
                    z = input("Are you sure you want to cancel all the tickets [y for yes and n for no] :\n")
                    if z == "y":
                        for i in range(1,n+1):
                            q11 = "update ticket_info set Status = %s where PNR = %s and passenger_no = %s"
                            l1 = ["Cancelled",P,i]
                            cur.execute(q11,l1)
                    print("All the tickets have been cancelled successfully")
                else:
                    for i in range(1,n+1):
                        v = int(input(f"Enter the passenger no whose ticket you want to cancel :\n"))
                        q11 = "update ticket_info set Status = %s where PNR = %s and passenger_no = %s"
                        l1 = ["Cancelled",P,v]
                        cur.execute(q11,l1)
                    print("Your ticket has been cancelled successfully")
            else:
                print("Your ticket is already cancelled")
    # Function to say thank you to the user

    def Thankyou(self):
        print("Thank you for using our service")

a = IRCTC()
a.Welcome()

c = False

a.Signin()

# if user is signed in successfully then only he/she can book a ticket or search a ticket or cancel a ticket

if c == True:

    h = int(input("\nWhat are you up to ? [1 for Booking a ticket, 2 for Searching your ticket using PNR, 3 Cancellation of a ticket of any person] :\n"))
    
    if h == 1:
        a.Booking()
        a.Thankyou()
    elif h == 2:
        a.Searching()
        a.Thankyou()
    else:
        a.cancellation()
        a.Thankyou()
    
    mydb.commit()
