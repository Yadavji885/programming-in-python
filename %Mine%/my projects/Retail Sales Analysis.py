import openpyxl
from openpyxl import Workbook

workbook = openpyxl.load_workbook("dataset.xlsx")
d = workbook["dataset"]

namew = input("Enter the name you want to give to the new file with sorted data :\n")
names = input("Enter the name you want to give to sheet :\n")

w = Workbook()
sheet = w.active
sheet.title = f"{names}"
w.save(f"{namew}.xlsx")

workbook1 = openpyxl.load_workbook(f"{namew}.xlsx")
o = workbook1[f"{names}"]

class Retail_sales_Analysis:

    def Organising(self):
        Months = []
        Companies = []
        Products = []
        rows = d.max_row
        cols = d.max_column

        mon_ord = ["January","February","March","April","May","June","July","August","September","October","November","December"]
        
        for i in range (3,rows+1):

            Months += [d.cell(i,1).value]

            Companies += [d.cell(i,2).value]

            Products += [d.cell(i,3).value]

        Months = set(Months)
        Companies = set(Companies)
        Products = set(Products)

        m = list(Months)
        c = list(Companies)
        p = list(Products)

        Months = []
        c.sort()

        for j in mon_ord:

            if j in m:

                Months.append(j)

        z = 1
        a = 0
        y = 1
        comp = []

        for x in range(0,len(Months)):
            
            for k in range(3,rows+1):

                if d.cell(k,1).value == Months[x]:

                    a += 1

                    comp.append(d.cell(k,2).value)

            for b in range(z,a+1):

                o.cell(b,1).value = Months[x]

            comp.sort()

            i = 0

            for q in range(z,a+1):

                o.cell(q,2).value = comp[i]

                if q == a+1:

                    for z in range(i,(len(comp)+1)):

                        o.cell(q,2).value == comp[z]

                i += 1

            for p in range(len(comp)):

                for u in range(1,rows+1):

                    if d.cell(u,1).value == Months[x]:

                        if d.cell(u,2).value == comp[p]:

                            o.cell(y,3).value = d.cell(u,3).value
                            o.cell(y,4).value = d.cell(u,4).value
                            o.cell(y,5).value = d.cell(u,5).value
                            d.cell(u,2).value = 0

                            y += 1
            
            y = a+1
            z = a+1
            i = 0
            comp.clear()

        for u in range(2):

            o.insert_rows(1)

        con = []

        for y in range(1,cols+1):

            con += [d.cell(1,y).value]

        for t in range(1,cols+1):
            
            o.cell(1,t).value = con[t-1]

a = Retail_sales_Analysis()
a.Organising()
workbook1.save(f"{namew}.xlsx")